import os
import sys
import io
import json
import pandas as pd
import openpyxl
from typing import List, Dict, Any, Optional
from pydantic import BaseModel, Field

from llm.base import BaseLLM
from information_extractor.base import BaseIE, ExtractionResult

class DataRange(BaseModel):
    sheet_name: str = Field(description="Tên của sheet chứa vùng dữ liệu này.")
    select_cols: List[str] = Field(description="Danh sách các ký tự cột cần lấy, ví dụ: ['A', 'C', 'D'].")
    start_row: int = Field(description="Dòng bắt đầu chứa dữ liệu (dòng data đầu tiên sau tiêu đề).")
    end_row: int = Field(description="Dòng kết thúc của vùng dữ liệu cần trích xuất.")

class SpreadsheetLayout(BaseModel):
    ranges: List[DataRange] = Field(default=[], description="Danh sách các vùng dữ liệu tiềm năng được tìm thấy trên sheet.")

class ExcelIE(BaseIE):
    def __init__(self, llm: BaseLLM, coder_llm: BaseLLM, temp_dir: str = "processed_files/temp"):
        self.llm = llm            
        self.coder_llm = coder_llm  
        self.temp_dir = temp_dir
        os.makedirs(self.temp_dir, exist_ok=True)
    
    def _select_range(self, query: str, processed_sheet_image: str) -> SpreadsheetLayout:
        prompt = f"""You are an expert data analyst looking at a compressed screenshot of an Excel sheet.
Based on the user's data request, locate the relevant data ranges (columns and rows) inside this sheet.

User Request: "{query}"

Instructions:
1. Identify which columns (e.g., A, B, C) and rows (e.g., from row 2 to 294) contain the target tables.
2. Even if the data is collapsed with '...', infer the boundary based on the row indices shown on the left.

Output the coordinates strictly matching the requested JSON schema. If the sheet contains no relevant data at all, return an empty ranges list: {{"ranges": []}}"""

        try:
            layout: SpreadsheetLayout = self.llm.generate(
                prompt=prompt,
                files=[processed_sheet_image],
                OutputModel=SpreadsheetLayout
            )
            return layout
        except Exception as e:
            # SỬA LỖI PYDANTIC: Nếu LLM trả về JSON lỗi hoặc {} không bóc tách được, trả về đối tượng rỗng an toàn
            print(f"   ⚠️ [DEBUG_WARNING] LLM không thể parse cấu trúc Layout hợp lệ hoặc trả về rỗng. Chi tiết: {e}")
            return SpreadsheetLayout(ranges=[])
    
    def _create_csv(self, layout: SpreadsheetLayout, original_file_path: str, sheet_idx: int) -> str:
        wb = openpyxl.load_workbook(original_file_path, data_only=True)
        
        sheet_names = wb.sheetnames
        if sheet_idx >= len(sheet_names):
            print(f"   ❌ [DEBUG_ERROR] Sheet Index {sheet_idx} vượt quá số lượng sheet của file Excel gốc.")
            return ""
        ws = wb[sheet_names[sheet_idx]]
        
        all_sliced_dfs = []
        
        for r_idx, r in enumerate(layout.ranges):
            print(f"   [DEBUG_CSV] Đang cắt Range #{r_idx}: Columns {r.select_cols} | Rows {r.start_row} -> {r.end_row}")
            
            # --- SỬA LỖI MẤT TIÊU ĐỀ (HEADER): Đọc tiêu đề gốc ở dòng 1 trước ---
            headers = []
            for col_letter in r.select_cols:
                header_val = ws[f"{col_letter}1"].value
                # Nếu tiêu đề ô bị trống, tự đặt tên cột tạm thời theo ký tự cột
                headers.append(header_val if header_val is not None else f"Column_{col_letter}")
            
            # Đọc dữ liệu data rows (từ start_row đến end_row)
            data = []
            for row_idx in range(r.start_row, r.end_row + 1):
                row_data = []
                for col_letter in r.select_cols:
                    cell_value = ws[f"{col_letter}{row_idx}"].value
                    row_data.append(cell_value)
                data.append(row_data)
                
            if data:
                # Tạo DataFrame sử dụng danh sách Headers dòng 1 làm tiêu đề chuẩn xác
                df_slice = pd.DataFrame(data, columns=headers)
                all_sliced_dfs.append(df_slice)
                
        if not all_sliced_dfs:
            print("   ⚠️ [DEBUG_CSV] Không có dữ liệu nào được bóc tách từ các ranges chỉ định.")
            return ""
            
        final_df = pd.concat(all_sliced_dfs, ignore_index=True)
        csv_path = os.path.join(self.temp_dir, f"temp_sheet_{sheet_idx}.csv")
        final_df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"   💾 [DEBUG_CSV] Đã tạo file CSV thành công: {csv_path} (Hình dáng: {final_df.shape})")
        return csv_path
        
    def _execute_code(self, code: str, csv_path: str) -> str:
        old_stdout = sys.stdout
        redirected_output = io.StringIO()
        sys.stdout = redirected_output
        
        local_vars = {"csv_path": csv_path}
        
        try:
            exec(code, {}, local_vars)
        except Exception as e:
            print(f"Lỗi thực thi Code nội bộ: {str(e)}")
        finally:
            sys.stdout = old_stdout
            
        return redirected_output.getvalue()
    
    def extract(self, query: str, processed_file_paths: List[str], original_file_path: str) -> ExtractionResult:
        sheet_results = []
        
        print(f"\n[EXCEL_IE_START] Phân tích file: {original_file_path}")
        print(f"Total processed sheet images found: {len(processed_file_paths)}")
        
        for idx, sheet_img_path in enumerate(processed_file_paths):
            print(f"\n--- 📑 Đang xử lý Sheet Index: {idx} ---")
            print(f"   ↳ Ảnh đầu vào: {sheet_img_path}")
            
            # Bước A: Định vị vùng dữ liệu
            layout = self._select_range(query, sheet_img_path)
            print(f"   📊 Layout nhận diện được: {layout.model_dump_json()}")
            
            if not layout.ranges:
                print(f"   ⏭️ [SKIPPED] Sheet {idx} không chứa vùng dữ liệu thích hợp.")
                continue
                
            # Bước B: Trích xuất vùng đó từ file gốc ra file CSV phẳng
            csv_path = self._create_csv(layout, original_file_path, idx)
            if not csv_path or not os.path.exists(csv_path):
                print(f"   ⏭️ [SKIPPED] Không tạo được CSV tạm thời cho Sheet {idx}.")
                continue
                
            # Đọc nhanh CSV để cấp Schema đầy đủ tiêu đề cho Coder LLM
            df_temp = pd.read_csv(csv_path)
            columns_schema = {str(col): str(dtype) for col, dtype in df_temp.dtypes.items()}
            sample_data = df_temp.head(3).to_dict(orient="records") # Lấy hẳn 3 dòng mẫu nhìn cho rõ tiêu đề dữ liệu
            
            print(f"   📋 Tiêu đề CSV thực tế thu được: {list(columns_schema.keys())}")
            
            # Bước C: Gọi Coder LLM
            code_prompt = f"""You are an expert Python Data Scientist. Your task is to write a short, clean Python script to answer the user's query by analyzing a CSV file.

User Query: "{query}"

The CSV file is located at the path stored in the variable `csv_path`.
CSV Columns & Types (Guaranteed to have original headers): {columns_schema}
CSV Sample Data (First 3 rows): {sample_data}

CRITICAL INSTRUCTIONS:
1. Use `pandas` to read the file via `pd.read_csv(csv_path)`.
2. Perform all necessary filtering, aggregation, or calculation to answer the query accurately.
3. You MUST use `print()` to print out the final answer clearly.
4. Return ONLY valid executable Python code block. Do not include markdown code ticks (```python) inside your actual execution path.

Write the Python code now:"""

            code_response = self.coder_llm.generate(prompt=code_prompt, files=None)
            raw_code = code_response.text.replace("```python", "").replace("```", "").strip()
            
            print(f"   💻 [DEBUG_CODE] Đoạn code sinh ra:\n{'-'*30}\n{raw_code}\n{'-'*30}")
            
            # Bước D: Chạy code và lấy kết quả đầu ra
            execution_output = self._execute_code(raw_code, csv_path)
            print(f"   🖥️ [DEBUG_STDOUT] Kết quả thực thi:\n{execution_output.strip()}")
            
            sheet_results.append(f"--- Kết quả từ Sheet {idx} ---\n{execution_output}")
            
            # Tạm thời comment xóa file để bạn có thể mở thư mục temp kiểm tra thủ công dữ liệu csv
            # if os.path.exists(csv_path):
            #     os.remove(csv_path)

        if not sheet_results:
            print("\n❌ [EXCEL_IE_END] Không thu thập được kết quả nào từ tất cả các sheets.")
            return ExtractionResult(status="NOT_FOUND", extracted_info="Không tìm thấy dữ liệu phù hợp trên các sheet.", quotes=[], missing_info=None)

        # 2. Tổng hợp kết quả từ các sheet
        all_outputs_text = "\n".join(sheet_results)
        
        synthesis_prompt = f"""You are a data verifier. Analyze the calculated outputs from multiple excel sheets and format them into the required response schema.

User Original Query: "{query}"
Calculated Outputs from Code Execution:
{all_outputs_text}

Fill out the schema. If the calculations fully answer the query, set status to 'FOUND'."""

        print("\n🧠 Đang tổng hợp kết quả tính toán cuối cùng về định dạng ExtractionResult...")
        final_result: ExtractionResult = self.llm.generate(
            prompt=synthesis_prompt,
            files=None,
            OutputModel=ExtractionResult
        )
        return final_result