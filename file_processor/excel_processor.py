import os
import html
import datetime
from dataclasses import dataclass, field
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright


@dataclass
class GridCell:
    value: Any = None
    row_span: int = 1
    col_span: int = 1
    is_covered: bool = False  # ô nằm trong vùng merge nhưng không phải ô neo (anchor)


class ExcelProcessor:
    def __init__(self, compress_threshold: int = 3, max_col_render: Optional[int] = None):
        self.compress_threshold = compress_threshold
        self.max_col_render = max_col_render

    def _load_sheet_grid(self, ws) -> list[list[GridCell]]:
        max_row = ws.max_row
        max_col = ws.max_column
        if self.max_col_render:
            max_col = min(max_col, self.max_col_render)

        grid = [[GridCell() for _ in range(max_col)] for _ in range(max_row)]

        covered = set()
        anchor_span = {}
        for mr in ws.merged_cells.ranges:
            r0, c0, r1, c1 = mr.min_row, mr.min_col, mr.max_row, mr.max_col
            if c0 > max_col:
                continue
            c1 = min(c1, max_col)
            anchor_span[(r0, c0)] = (r1 - r0 + 1, c1 - c0 + 1)
            for r in range(r0, r1 + 1):
                for c in range(c0, c1 + 1):
                    if (r, c) != (r0, c0):
                        covered.add((r, c))

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                gc = grid[r - 1][c - 1]
                if (r, c) in covered:
                    gc.is_covered = True
                    continue
                cell = ws.cell(row=r, column=c)
                gc.value = cell.value
                if (r, c) in anchor_span:
                    gc.row_span, gc.col_span = anchor_span[(r, c)]
        return grid

    @staticmethod
    def _cell_type(value: Any) -> str:
        if value is None or value == "":
            return "empty"
        if isinstance(value, bool):
            return "bool"
        if isinstance(value, (int, float)):
            return "num"
        if isinstance(value, (datetime.date, datetime.datetime)):
            return "date"
        if isinstance(value, str):
            return "str"
        return "other"

    def _row_signature(self, row: list[GridCell]) -> tuple:
        sig = []
        for gc in row:
            if gc.is_covered:
                sig.append("covered")
            else:
                sig.append(self._cell_type(gc.value))
        return tuple(sig)

    def _row_blocks_merge(self, row: list[GridCell]) -> bool:
        return any(gc.row_span > 1 or gc.is_covered for gc in row)

    def _find_compress_runs(self, grid: list[list[GridCell]]) -> list[tuple[int, int]]:
        n = len(grid)
        signatures = [self._row_signature(r) for r in grid]
        blocked = [self._row_blocks_merge(r) for r in grid]

        runs = []
        i = 0
        while i < n:
            if blocked[i]:
                i += 1
                continue
            j = i
            while (j + 1 < n) and (not blocked[j + 1]) and (signatures[j + 1] == signatures[i]):
                j += 1
            if (j - i + 1) > self.compress_threshold:
                runs.append((i, j))
            i = j + 1
        return runs

    def _compress_grid(self, grid: list[list[GridCell]]):
        runs = self._find_compress_runs(grid)
        run_start = {i: j for i, j in runs}

        compressed = []  # list of (row_label, list[GridCell])
        n_cols = len(grid[0]) if grid else 0
        i = 0
        n = len(grid)
        n_hidden_total = 0
        while i < n:
            if i in run_start:
                j = run_start[i]
                compressed.append((str(i + 1), grid[i]))
                ellipsis_row = [GridCell(value="...") for _ in range(n_cols)]
                compressed.append(("...", ellipsis_row))
                compressed.append((str(j + 1), grid[j]))
                n_hidden_total += (j - i + 1) - 2
                i = j + 1
            else:
                compressed.append((str(i + 1), grid[i]))
                i += 1
        return compressed, n_hidden_total

    def _render_html(self, compressed, n_cols: int, sheet_name: str = "") -> str:
        col_letters = [get_column_letter(c) for c in range(1, n_cols + 1)]
        style = """
        <style>
          body { font-family: Arial, Helvetica, sans-serif; margin: 0; padding: 16px; background: #fff; }
          table { border-collapse: collapse; }
          td, th {
            border: 1px solid #b7b7b7; padding: 4px 12px; font-size: 14px;
            text-align: center; white-space: nowrap; height: 22px;
          }
          th { background: #d9d9d9; font-weight: bold; color: #333; }
          td.row-idx { background: #d9d9d9; font-weight: bold; color: #333; }
          td.ellipsis { color: #888; font-weight: bold; }
          td.text-cell { text-align: left; }
          .title { font-family: Arial; font-size: 13px; color: #555; margin-bottom: 6px; }
        </style>
        """
        rows_html = []
        header_cells = "<th></th>" + "".join(f"<th>{c}</th>" for c in col_letters)
        rows_html.append(f"<tr>{header_cells}</tr>")

        for row_label, row in compressed:
            cells_html = f'<td class="row-idx">{html.escape(row_label)}</td>'
            for gc in row:
                if gc.is_covered:
                    continue
                attrs = ""
                if gc.row_span > 1:
                    attrs += f' rowspan="{gc.row_span}"'
                if gc.col_span > 1:
                    attrs += f' colspan="{gc.col_span}"'
                val = "" if gc.value is None else str(gc.value)
                is_ellipsis = val == "..."
                cls = "ellipsis" if is_ellipsis else ("text-cell" if isinstance(gc.value, str) else "")
                cells_html += f'<td class="{cls}"{attrs}>{html.escape(val)}</td>'
            rows_html.append(f"<tr>{cells_html}</tr>")

        table_html = "<table>" + "".join(rows_html) + "</table>"
        title = f'<div class="title">Sheet: {html.escape(sheet_name)}</div>' if sheet_name else ""
        return f"<html><head>{style}</head><body>{title}{table_html}</body></html>"

    def _html_to_image(self, html_str: str, output_path: str):
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            page.set_content(html_str)
            page.screenshot(path=output_path, full_page=True)
            browser.close()

    def vertical_compress_excel_to_image(
        self,
        excel_file_path: str,
        output_dir: Optional[str] = None,
        sheet_names: Optional[list[str]] = None,
    ) -> list[dict]:
        """
        Return: list dict, mỗi phần tử ứng với 1 sheet:
            {
                "sheet": <tên sheet>,
                "image_path": <đường dẫn ảnh>,
                "original_rows": <số dòng gốc>,
                "rendered_rows": <số dòng sau khi nén>,
                "hidden_rows": <số dòng đã bị ẩn/nén>,
            }
        """
        wb = openpyxl.load_workbook(excel_file_path, data_only=True)
        sheets = [wb[s] for s in sheet_names] if sheet_names else wb.worksheets

        output_dir = output_dir or os.path.dirname(os.path.abspath(excel_file_path))
        os.makedirs(output_dir, exist_ok=True)
        base = os.path.splitext(os.path.basename(excel_file_path))[0]

        results = []
        for ws in sheets:
            grid = self._load_sheet_grid(ws)
            if not grid:
                continue
            n_cols = len(grid[0])
            compressed, n_hidden = self._compress_grid(grid)
            html_str = self._render_html(compressed, n_cols, ws.title)

            safe_sheet = "".join(c if c.isalnum() else "_" for c in ws.title)
            out_path = os.path.join(output_dir, f"{base}__{safe_sheet}__compressed.png")
            self._html_to_image(html_str, out_path)

            results.append({
                "sheet": ws.title,
                "image_path": out_path,
                "original_rows": len(grid),
                "rendered_rows": len(compressed),
                "hidden_rows": n_hidden,
            })
        return results


if __name__ == "__main__":
    processor = ExcelProcessor(compress_threshold=3)
    path = "climateMeasurements.xlsx"
    out = processor.vertical_compress_excel_to_image(path, output_dir="out")
    print(out)

